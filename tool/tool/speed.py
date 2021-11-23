"""
    テンプレ諸元選択
"""
# coding=utf-8
import os
import sys
import json
import csv
import glob
import openpyxl

"""
    初期設定
"""
master_json = {
    'key': 'value'
}

os.chdir(r"/Applications/MAMP/htdocs/python/tool/dist")

def main(area):
    os.chdir(r"/Applications/MAMP/htdocs/python/tool/dist")

    # マスタ読み込み
    wb = openpyxl.load_workbook('master/テンプレートリスト_東京.xlsx')
    ws = wb['マスタ']
    items = []
    rowCount = 2
    for row in ws.iter_rows(min_row=2):
        key = ''
        value = ws.cell(row=rowCount, column=2).value
        for cell in row[2:]:
            key += str(cell.value)
        master_json[key] = value
        rowCount += 1
    # print(master_json)

    paths = glob.glob("csv/before/*.csv")
    files = []
    for path in paths:
        files.append(os.path.basename(path))

    # CSV読み込み
    for fileName in files:
        
        csvheader = []
        csvList = []
        # print(fileName)
        with open('csv/before/{}'.format(fileName), mode="r", encoding="utf-8") as rf:
            reader = csv.reader(rf)
            next(reader)
            for rows in reader:
                item = ''
                for col in rows[4:]:
                    item += col
                # csvList.append(master_json[item])
                if item in master_json:
                    csvList.append(master_json[item])
                else:
                    csvList.append('')

            # print(csvList)
            with open('csv/after/{}'.format(fileName), 'w') as wf:
                writer = csv.writer(wf)
                writer.writerow(['サイトID', 'オーダーID', 'テンプレ諸元', '優先諸元'])
                rf.seek(0)
                reader = csv.reader(rf)
                next(reader)
                for index, rows in enumerate(reader):
                    writer.writerow([rows[0], rows[1], csvList[index-1],''])

if __name__ == '__main__':
    print('---■ 処理開始 ■---')
    area = input('地域を入力してください。(例：東京))：')
    main(area)
    print('---■ 処理終了 ■---')
    # input("何かキーを押すと終了します。")

